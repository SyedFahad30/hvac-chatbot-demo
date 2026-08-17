"""
hvac-lambda_function.py
------------------------------------------------------------------
Backend for the Clermont Air & Heat (placeholder HVAC business) chat
widget AND missed-call text-back flow. Same architecture as the Lake
America build, extended with two Twilio webhook routes.

Three entry points, all through this one Lambda (routed by rawPath):

  POST /            - website chat widget (JSON body, same as before)
  GET  /            - staff dashboard reads current job-request data
  POST /voice        - Twilio Voice webhook. A call only reaches this
                        number after the real business line rang unanswered
                        and the business's phone carrier forwarded it here
                        (see README for how that forwarding is set up).
                        Responds with TwiML that plays a short message and
                        hangs up, then fires an outbound SMS to the caller.
  POST /sms          - Twilio SMS webhook. Every reply the caller texts
                        back comes through here. Conversation state is
                        keyed by phone number (not a browser session,
                        since there's no session concept over SMS) and
                        replies use the same Claude call as the widget.

ENVIRONMENT VARIABLES:
  ANTHROPIC_API_KEY     - Anthropic API key (server-side only)
  DYNAMODB_TABLE        - conversations table name
  S3_BUCKET             - bucket holding the job-intake workbook
  S3_KEY                - object key, e.g. new-job-intake.xlsx
  ALLOWED_ORIGIN         - CORS origin for the website widget
  TWILIO_ACCOUNT_SID    - from twilio.com console (leave blank until you
                           have an account — /voice and /sms will no-op
                           safely if unset)
  TWILIO_AUTH_TOKEN     - from twilio.com console
  TWILIO_PHONE_NUMBER   - the Twilio number itself, e.g. +14075551234
"""

import json
import os
import re
import time
import uuid
import io
import traceback
import base64
import urllib.request
import urllib.error
import urllib.parse

import boto3
from openpyxl import Workbook, load_workbook

dynamodb = boto3.resource("dynamodb")
s3 = boto3.client("s3")

TABLE_NAME = os.environ.get("DYNAMODB_TABLE", "hvac-conversations")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
S3_BUCKET = os.environ.get("S3_BUCKET", "")
S3_KEY = os.environ.get("S3_KEY", "new-job-intake.xlsx")
ALLOWED_ORIGIN = os.environ.get("ALLOWED_ORIGIN", "*")

TWILIO_ACCOUNT_SID = os.environ.get("TWILIO_ACCOUNT_SID", "")
TWILIO_AUTH_TOKEN = os.environ.get("TWILIO_AUTH_TOKEN", "")
TWILIO_PHONE_NUMBER = os.environ.get("TWILIO_PHONE_NUMBER", "")

ANTHROPIC_URL = "https://api.anthropic.com/v1/messages"
ANTHROPIC_MODEL = "claude-sonnet-4-6"

BUSINESS_NAME = "Clermont Air & Heat"

# Placeholder service area / pricing facts — swap for the real business's
# details before using this with an actual client.
SYSTEM_PROMPT = f"""You are the dispatch/booking assistant for {BUSINESS_NAME}, a residential HVAC company serving Clermont, FL and the surrounding Lake County area.

Real facts about the business (PLACEHOLDER DATA — replace before real use):
- Hours: Monday-Saturday 7am-7pm. Emergency/after-hours service available for an additional fee.
- Services: AC repair, AC installation/replacement, heating repair, furnace/heat pump service, seasonal maintenance plans, duct cleaning, free replacement estimates.
- Service area: Clermont, Minneola, Winter Garden, Groveland, and surrounding Lake County FL.
- Typical response time: same-day for emergencies, 1-2 business days for routine service.
- Financing available for full system replacements.

Your job:
1. Figure out urgency first: is this an emergency (no AC/heat, active leak, burning smell, no power to unit) or routine (maintenance, estimate, non-urgent repair)? Say so back to the customer so they know what to expect.
2. For emergencies: reassure them, say a technician will call to confirm timing, and prioritize collecting name, phone, and address quickly — don't drag out small talk.
3. For routine requests: ask what's going on with their system, whether it's repair vs. new install/estimate, preferred days/times, then name, phone, and service address.
4. You do NOT have access to a live technician schedule. Never confirm a specific appointment time — say dispatch will call to confirm.
5. Give NO specific repair diagnosis or pricing quotes over chat — those require an in-person look. You can mention that estimates are free.
6. If someone describes a genuine safety hazard (gas smell, sparking, smoke, carbon monoxide alarm), tell them to leave the area and call 911 or the gas company immediately — do not continue troubleshooting with them.
7. Warm, direct, no-nonsense tone — this is a trade business, not a spa. 2-4 sentences per reply."""

PHONE_RE = re.compile(r"(\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}")

EXTRACTION_PROMPT = """Read the conversation transcript below between an HVAC company's dispatch assistant and a customer. Extract the job details as strict JSON with EXACTLY these seven keys, nothing more:

{
  "name": "the customer's full name, or empty string if not given",
  "phone": "the customer's phone number exactly as they typed it, or empty string",
  "address": "the service address if mentioned, or empty string",
  "urgency": "one of: emergency, routine",
  "issue": "short plain-text description under 12 words of what's wrong or requested",
  "preferred_time": "the day/date/time they asked for, in their own words, or empty string",
  "notes": "anything else relevant, under 15 words, or empty string"
}

Rules:
- Output ONLY the raw JSON object. No markdown code fences, no explanation, no leading/trailing text.
- Every value must be a plain string, never an object, array, or the raw transcript.
- If a field truly wasn't mentioned, use an empty string "" for it — never guess.

TRANSCRIPT:
"""

INTAKE_FIELDS = ["name", "phone", "address", "urgency", "issue", "preferred_time", "notes"]


def _cors_headers():
    return {
        "Access-Control-Allow-Origin": ALLOWED_ORIGIN,
        "Access-Control-Allow-Headers": "Content-Type",
        "Access-Control-Allow-Methods": "OPTIONS,POST,GET",
        "Content-Type": "application/json",
    }


def _response(status, body_dict):
    return {
        "statusCode": status,
        "headers": _cors_headers(),
        "body": json.dumps(body_dict),
    }


def _twiml_response(xml_body):
    return {
        "statusCode": 200,
        "headers": {"Content-Type": "text/xml"},
        "body": xml_body,
    }


def _xml_escape(text):
    return (
        (text or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _call_anthropic_raw(system_prompt, messages, max_tokens=300):
    payload = json.dumps(
        {
            "model": ANTHROPIC_MODEL,
            "max_tokens": max_tokens,
            "system": system_prompt,
            "messages": messages,
        }
    ).encode("utf-8")

    req = urllib.request.Request(
        ANTHROPIC_URL,
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=25) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8")
        raise RuntimeError(f"Anthropic API error {e.code}: {err_body}")

    for block in data.get("content", []):
        if block.get("type") == "text":
            return block["text"]
    return ""


def call_anthropic(messages):
    reply = _call_anthropic_raw(SYSTEM_PROMPT, messages, max_tokens=300)
    return reply or "Sorry, could you try that again?"


def _parse_intake_json(raw):
    if not raw:
        return None
    candidates = [raw.strip()]
    fenced = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.IGNORECASE)
    candidates.append(fenced.strip())
    match = re.search(r"\{.*\}", raw, flags=re.DOTALL)
    if match:
        candidates.append(match.group(0))
    for candidate in candidates:
        try:
            parsed = json.loads(candidate)
            if isinstance(parsed, dict):
                return parsed
        except (json.JSONDecodeError, TypeError):
            continue
    return None


def extract_intake_fields(messages):
    transcript = "\n".join(f"{m['role']}: {m['content']}" for m in messages)
    raw = _call_anthropic_raw(
        "You extract structured data from conversations and reply with ONLY a raw JSON object, nothing else — no markdown, no commentary.",
        [{"role": "user", "content": EXTRACTION_PROMPT + transcript}],
        max_tokens=300,
    )
    parsed = _parse_intake_json(raw)
    if parsed is None:
        print("INTAKE EXTRACTION: could not parse JSON from model output:")
        print(raw)
        parsed = {}
    clean = {}
    for field in INTAKE_FIELDS:
        value = parsed.get(field, "")
        clean[field] = value if isinstance(value, str) else ""
    return clean


def save_conversation(session_id, messages, lead_logged):
    table = dynamodb.Table(TABLE_NAME)
    table.put_item(
        Item={
            "sessionId": session_id,
            "messages": json.dumps(messages),
            "leadLogged": lead_logged,
            "updatedAt": int(time.time()),
            "ttl": int(time.time()) + 30 * 24 * 60 * 60,
        }
    )


def get_conversation_meta(session_id):
    table = dynamodb.Table(TABLE_NAME)
    resp = table.get_item(Key={"sessionId": session_id})
    return resp.get("Item")


def append_row_to_excel(row_values):
    if not S3_BUCKET:
        return
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=S3_KEY)
        wb = load_workbook(io.BytesIO(obj["Body"].read()))
        ws = wb.active
    except s3.exceptions.NoSuchKey:
        wb = Workbook()
        ws = wb.active
        ws.title = "New Job Intake"
        ws.append(["Timestamp", "Name", "Phone", "Address", "Urgency", "Issue", "Preferred Time", "Notes", "Channel", "Session ID"])
    ws.append(row_values)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    s3.put_object(
        Bucket=S3_BUCKET,
        Key=S3_KEY,
        Body=buf.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def read_dashboard_rows():
    if not S3_BUCKET:
        return []
    try:
        obj = s3.get_object(Bucket=S3_BUCKET, Key=S3_KEY)
        wb = load_workbook(io.BytesIO(obj["Body"].read()))
        ws = wb.active
    except s3.exceptions.NoSuchKey:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        record = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else ""
            record[header] = str(value) if value is not None else ""
        records.append(record)
    return records


def maybe_log_lead(session_id, messages, already_logged, channel="web"):
    if already_logged:
        return already_logged
    full_text = "\n".join(f"{m['role']}: {m['content']}" for m in messages)
    if not PHONE_RE.search(full_text) and channel == "web":
        # Over SMS, we already know the phone number from the caller ID, so
        # don't require it to appear in-text the way the web widget does.
        return already_logged

    fields = extract_intake_fields(messages)
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
    append_row_to_excel(
        [
            timestamp,
            fields.get("name", ""),
            fields.get("phone", ""),
            fields.get("address", ""),
            fields.get("urgency", ""),
            fields.get("issue", ""),
            fields.get("preferred_time", ""),
            fields.get("notes", ""),
            channel,
            session_id,
        ]
    )
    return True


# ------------------------------------------------------------------
# TWILIO HELPERS
# ------------------------------------------------------------------

def send_sms(to_number, body_text):
    """Sends an outbound SMS via Twilio's REST API directly (no twilio SDK
    dependency — just an HTTPS POST with basic auth, same minimal-dependency
    approach as the rest of this Lambda)."""
    if not (TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN and TWILIO_PHONE_NUMBER):
        print("Twilio not configured — skipping SMS send (this is expected until you have a Twilio account).")
        return

    url = f"https://api.twilio.com/2010-04-01/Accounts/{TWILIO_ACCOUNT_SID}/Messages.json"
    payload = urllib.parse.urlencode({
        "To": to_number,
        "From": TWILIO_PHONE_NUMBER,
        "Body": body_text,
    }).encode("utf-8")

    auth = base64.b64encode(f"{TWILIO_ACCOUNT_SID}:{TWILIO_AUTH_TOKEN}".encode()).decode()
    req = urllib.request.Request(
        url, data=payload, method="POST",
        headers={
            "Authorization": f"Basic {auth}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            resp.read()
    except urllib.error.HTTPError as e:
        print("TWILIO SEND FAILED:", e.read().decode("utf-8"))


def _parse_form_body(event):
    """Twilio webhooks send application/x-www-form-urlencoded bodies, not
    JSON. API Gateway may base64-encode the body depending on config."""
    body = event.get("body") or ""
    if event.get("isBase64Encoded"):
        body = base64.b64decode(body).decode("utf-8")
    return dict(urllib.parse.parse_qsl(body))


def handle_voice_webhook(event):
    """A call only reaches this route because the business's real phone
    line already rang unanswered and their carrier forwarded it here (see
    README — this is set up as conditional call forwarding on their end,
    not something Twilio or this Lambda controls). Instead of just hanging
    up, we ask the caller to describe their issue right on the call using
    <Gather input="speech">, which Twilio transcribes and POSTs to
    /voice-collect below — that's where the lead actually gets logged."""
    safe_name = _xml_escape(BUSINESS_NAME)
    twiml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        '<Gather input="speech" action="/voice-collect" method="POST" '
        'speechTimeout="auto" timeout="6" maxSpeechTime="60">'
        f"<Say>Thanks for calling {safe_name}. Please briefly describe the issue you're having, after the tone.</Say>"
        "</Gather>"
        "<Say>Sorry, we didn't catch that. We'll follow up with you shortly.</Say>"
        "<Hangup/>"
        "</Response>"
    )
    return _twiml_response(twiml)


def handle_voice_collect_webhook(event):
    """Twilio posts here after <Gather input="speech"> finishes, with the
    transcribed text in SpeechResult. Log the lead with whatever we got —
    caller ID always gives us the phone number even if speech recognition
    came back empty — then read back a confirmation and hang up."""
    form = _parse_form_body(event)
    caller = form.get("From", "")
    speech_result = form.get("SpeechResult", "").strip()

    urgency = "unknown"
    issue_text = speech_result if speech_result else "No description captured (caller didn't respond or wasn't understood)"

    if speech_result:
        try:
            classification = _call_anthropic_raw(
                "You classify HVAC phone issues from a short spoken description. "
                "Reply with ONLY one word: 'emergency' (no AC/heat, active leak, burning smell, "
                "no power to unit, safety hazard) or 'routine' (everything else).",
                [{"role": "user", "content": speech_result}],
                max_tokens=10,
            )
            classification_clean = classification.strip().lower()
            if "emergency" in classification_clean:
                urgency = "emergency"
            elif "routine" in classification_clean:
                urgency = "routine"
        except Exception:
            print("VOICE URGENCY CLASSIFY FAILED:")
            print(traceback.format_exc())

    try:
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
        append_row_to_excel(
            [
                timestamp,
                "",                # name - not asked for on the call
                caller,
                "",                # address - not asked for on the call
                urgency,
                issue_text[:200],
                "",                # preferred_time
                "Captured via voice call (missed-call flow)",
                "missed-call",
                f"call-{caller}-{int(time.time())}",
            ]
        )
    except Exception:
        print("VOICE LEAD LOGGING FAILED:")
        print(traceback.format_exc())

    safe_name = _xml_escape(BUSINESS_NAME)
    if speech_result:
        confirmation = f"Thanks, we've got that noted. We'll reach you at this number shortly to get you scheduled."
    else:
        confirmation = f"Sorry we couldn't hear you clearly. We've got your number and will reach out shortly."
    twiml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        "<Response>"
        f"<Say>{_xml_escape(confirmation)}</Say>"
        "<Hangup/>"
        "</Response>"
    )
    return _twiml_response(twiml)


def handle_sms_webhook(event):
    """Every reply the customer texts back comes through here. Conversation
    state is keyed by their phone number (there's no browser session over
    SMS), stored in the same DynamoDB table as the web widget uses."""
    form = _parse_form_body(event)
    caller = form.get("From", "")
    incoming_text = form.get("Body", "")

    if not caller:
        return _twiml_response('<?xml version="1.0" encoding="UTF-8"?><Response></Response>')

    session_id = f"sms-{caller}"
    existing = get_conversation_meta(session_id) or {}
    history = json.loads(existing.get("messages", "[]")) if existing.get("messages") else []
    already_logged = bool(existing.get("leadLogged", False))

    history.append({"role": "user", "content": incoming_text})

    try:
        reply_text = call_anthropic(history)
    except Exception:
        print("SMS ANTHROPIC CALL FAILED:")
        print(traceback.format_exc())
        reply_text = f"Sorry, something went wrong on our end — please call us directly."

    full_messages = history + [{"role": "assistant", "content": reply_text}]

    try:
        logged_now = maybe_log_lead(session_id, full_messages, already_logged, channel="sms")
    except Exception:
        print("SMS LEAD LOGGING FAILED:")
        print(traceback.format_exc())
        logged_now = already_logged

    try:
        save_conversation(session_id, full_messages, logged_now)
    except Exception:
        print("SMS DYNAMODB SAVE FAILED:")
        print(traceback.format_exc())

    # Replying via TwiML <Message> sends the text back automatically —
    # no separate send_sms() call needed for in-conversation replies.
    escaped_reply = (
        reply_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )
    twiml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        f"<Response><Message>{escaped_reply}</Message></Response>"
    )
    return _twiml_response(twiml)


def lambda_handler(event, context):
    method = event.get("requestContext", {}).get("http", {}).get("method", "POST")
    path = event.get("rawPath", "/")

    if method == "OPTIONS":
        return _response(200, {})

    if path == "/voice" and method == "POST":
        return handle_voice_webhook(event)

    if path == "/voice-collect" and method == "POST":
        return handle_voice_collect_webhook(event)

    if path == "/sms" and method == "POST":
        return handle_sms_webhook(event)

    if method == "GET":
        try:
            records = read_dashboard_rows()
            return _response(200, {"records": records})
        except Exception:
            print("DASHBOARD READ FAILED:")
            print(traceback.format_exc())
            return _response(502, {"error": "Could not read intake data"})

    # ---- Website chat widget (POST /) — same behavior as Lake America ----
    try:
        body = json.loads(event.get("body") or "{}")
    except json.JSONDecodeError:
        return _response(400, {"error": "Invalid JSON body"})

    messages = body.get("messages")
    session_id = body.get("sessionId") or str(uuid.uuid4())

    if not messages or not isinstance(messages, list):
        return _response(400, {"error": "messages[] is required"})

    if not ANTHROPIC_API_KEY:
        return _response(500, {"error": "Server misconfigured: missing ANTHROPIC_API_KEY"})

    try:
        reply_text = call_anthropic(messages)
    except Exception as e:
        return _response(502, {"error": f"Upstream AI call failed: {str(e)}"})

    full_messages = messages + [{"role": "assistant", "content": reply_text}]
    existing = get_conversation_meta(session_id) or {}
    already_logged = bool(existing.get("leadLogged", False))

    try:
        logged_now = maybe_log_lead(session_id, full_messages, already_logged, channel="web")
    except Exception:
        print("LEAD LOGGING FAILED:")
        print(traceback.format_exc())
        logged_now = already_logged

    try:
        save_conversation(session_id, full_messages, logged_now)
    except Exception:
        print("DYNAMODB SAVE FAILED:")
        print(traceback.format_exc())

    return _response(200, {"reply": reply_text, "sessionId": session_id})
